from django.shortcuts import render,redirect
from django.contrib import messages
from . models import *
import random 
from django.db.models import Sum, Count
from django.db import connection
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.conf import settings
from datetime import datetime, timedelta
from datetime import date 
import datetime
import openpyxl
import os
import pickle

import pickle
import os

try:
    module_dir = os.path.dirname(__file__)
    file_path = os.path.join(module_dir, "kmeans_model_st.pkl")
    with open(file_path, "rb") as file:
        model = pickle.load(file)
except EOFError:
    print("EOFError: Ran out of input. The file may be corrupted or empty.")
except FileNotFoundError:
    print(f"File not found: {file_path}")
except Exception as e:
    print(f"An error occurred: {e}")


def student_login(request):
	if request.session.has_key('user_id'):
		return render(request,'dashboard.html',{})
	else:
		if request.method == 'POST':
			name=request.POST.get('username')
			pwd=request.POST.get('password')
			user_exist=Student_Detail.objects.filter(username=name,password=pwd)
			if user_exist:
				request.session['name']= request.POST.get('username')
				a = request.session['name']
				sess = Student_Detail.objects.only('id').get(username=a).id
				request.session['user_id']= sess
				return redirect('dashboard')
			else:
				messages.success(request,'Invalid username or Password')
		return render(request,'student_login.html',{})
def dashboard(request):
	return render(request,'dashboard.html',{})
def register(request):
	if request.method == 'POST':
		Name = request.POST.get('uname')
		Adddress = request.POST.get('address')
		Mobile= request.POST.get('mobile')
		Email = request.POST.get('email')
		Password = request.POST.get('pwd')
		unum = request.POST.get('username')
		country = request.POST.get('country')
		city = request.POST.get('city')
		state = request.POST.get('state')
		dob = request.POST.get('dob')
		gender = request.POST.get('gender')
		education = request.POST.get('education')
		cutoff_mark = request.POST.get('cutoff_mark')
		score = request.POST.get('score')
		image = request.FILES['image']
		caste =request.POST.get('caste')
		student_exist = Student_Detail.objects.filter(username=unum)
		if student_exist:
			messages.success(request,'Username No Already Exsit')
		else:
			crt = Student_Detail.objects.create(student_name=Name,
			address=Adddress,phone_number=Mobile,password=Password,email_id=Email,username=unum,country=country,
			city=city,state=state,dob=dob,caste=caste,gender=gender,education=education,Score=score,image=image,cutoff_mark=cutoff_mark)
			if crt:
				messages.success(request,'Registered Successfully')
	return render(request,'register.html',{})
def logout(request):
    try:
        del request.session['user_id']
        del request.session['name']
    except:
     pass
    return render(request, 'student_login.html', {})
from django.core.exceptions import ValidationError
def admin_home(request):
	if request.method == 'POST':
		file = request.FILES['excel_file']
		
		# Validate file type
		if not file.name.endswith('.xlsx'):
			messages.error(request, "Invalid file type. Please upload an Excel file.")
			return render(request, 'home.html', {})

		try:
			workbook = openpyxl.load_workbook(file, read_only=True)
		except Exception as e:
			messages.error(request, f"Error processing file: {str(e)}")
			return render(request, 'home.html', {})
		
		# Get name of the first sheet and then open sheet by name
		first_sheet = workbook.sheetnames[0]
		worksheet = workbook[first_sheet]
		
		data = []
		for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row): 
			detail = Engineering_College_Detail()
			detail.college_name = row[0].value
			detail.degree = row[1].value
			detail.department = row[2].value
			detail.OC = row[3].value
			detail.BC = row[4].value
			detail.BCM = row[5].value
			detail.MBCV = row[6].value
			detail.MBCDNC = row[7].value
			detail.MBA = row[8].value
			detail.SC = row[9].value
			detail.SCA = row[10].value
			detail.ST = row[11].value
			detail.skill = row[12].value
			data.append(detail)
		
		# Bulk create data
		crt = Engineering_College_Detail.objects.bulk_create(data)
		if crt:
			messages.success(request, "Detail Added Successfully.")
		else:
			messages.error(request, "Failed to add details.")
		
	return render(request, 'home.html', {})


def engineering_search(request):
	user_id=request.session['user_id']
	cur=connection.cursor()
	user = '''SELECT s.caste from app_student_detail as s where s.id='%d' ''' %(int(user_id))
	sel = cur.execute(user)
	user_detail = cur.fetchone()
	caste = user_detail[0]
	cursor=connection.cursor()
	sql=''' SELECT e.skill from app_engineering_college_detail as e GROUP BY e.skill'''
	post=cursor.execute(sql)
	row=cursor.fetchall()
	cursor1=connection.cursor()
	sql1=''' SELECT e.department from app_engineering_college_detail as e GROUP BY e.department'''
	post1=cursor1.execute(sql1)
	row1=cursor1.fetchall()
	if request.method == 'POST':
		degree=request.POST.get('degree')
		department=request.POST.get('department')
		cutoff_mark =request.POST.get('cutoff_mark')
		if caste == 'OC':
			a=Engineering_College_Detail.objects.filter(skill=degree,OC__lte=cutoff_mark)
			return render(request,'engineering_search.html',{'row':row,'a':a})
		elif caste == 'BC':
			a=Engineering_College_Detail.objects.filter(skill=degree,BC__lte=cutoff_mark)
			return render(request,'engineering_search.html',{'row':row,'a':a})
		elif caste == 'BCM':
			a=Engineering_College_Detail.objects.filter(skill=degree,BCM__lte=cutoff_mark)
			return render(request,'engineering_search.html',{'row':row,'a':a})
		elif caste == 'MBCV':
			a=Engineering_College_Detail.objects.filter(skill=degree,MBCV__lte=cutoff_mark)
			return render(request,'engineering_search.html',{'row':row,'a':a})
		elif caste == 'MBCDNC':
			a=Engineering_College_Detail.objects.filter(skill=degree,MBCDNC__lte=cutoff_mark)
			return render(request,'engineering_search.html',{'row':row,'a':a})
		elif caste == 'MBA':
			a=Engineering_College_Detail.objects.filter(skill=degree,MBA__lte=cutoff_mark)
			return render(request,'engineering_search.html',{'row':row,'a':a})
		elif caste == 'ST':
			a=Engineering_College_Detail.objects.filter(skill=degree,ST__lte=cutoff_mark)
			return render(request,'engineering_search.html',{'row':row,'a':a})
		elif caste == 'SC':
			a=Engineering_College_Detail.objects.filter(skill=degree,SC__lte=cutoff_mark)
			return render(request,'engineering_search.html',{'row':row,'a':a})	
		elif caste == 'SCA':
			a=Engineering_College_Detail.objects.filter(skill=degree,SCA__lte=cutoff_mark)
			return render(request,'engineering_search.html',{'row':row,'a':a})
	return render(request,'engineering_search.html',{'row':row,'row1':row1,'caste':caste})
def train_data(request):
	import pandas as pd
	from sklearn.preprocessing import StandardScaler
	from sklearn.cluster import KMeans
	from sklearn.impute import SimpleImputer
	current_file_path = os.path.dirname(os.path.abspath(__file__))
	app_folder_path = os.path.dirname(current_file_path)
	df = pd.read_csv(app_folder_path+'/dataset_new.csv')

	# Select the 'ST' column
	cutoff_data = df[['skill']]
	imputer = SimpleImputer(strategy='mean')
	cutoff_data_imputed = imputer.fit_transform(cutoff_data)
	scaler = StandardScaler()
	scaled_data = scaler.fit_transform(cutoff_data_imputed)

	# Train the K-means model
	k = 3  # Example: 3 clusters
	kmeans = KMeans(n_clusters=k, random_state=42)
	kmeans.fit(scaled_data)

	# Add the cluster labels to the original dataframe
	df['Cluster'] = kmeans.labels_

	# Display the dataframe with the cluster labels
	print(df[['address', 'skill', 'Cluster']])
	df['Cluster'] = kmeans.labels_

	# Display the dataframe with the cluster labels
	print(df[['address', 'skill', 'Cluster']])
	# Example: Recommend colleges from a specific cluster
	cluster_number = 0  # Specify the cluster number you want to recommend
	recommended_colleges = df[df['Cluster'] == cluster_number]

	print("Recommended Colleges and Course based on Skill :")
	print(recommended_colleges[['address', 'skill', 'Cluster']])
	import joblib
	# Save the K-means model
	joblib.dump(kmeans, 'kmeans_model_st.pkl')